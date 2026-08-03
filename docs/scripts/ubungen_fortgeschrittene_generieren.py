#!/usr/bin/env python3
"""Generate XLSX exercise files for Excel-lenz advanced course."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT_DIR = os.path.join(os.path.dirname(__file__), "../exercises/fortgeschrittene")
os.makedirs(OUT_DIR, exist_ok=True)

HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
INSTR_FILL = PatternFill('solid', fgColor='F2F2F2')
INSTR_FONT = Font(name='Calibri', italic=True, color='333333', size=10)
DATA_FONT = Font(name='Calibri', size=11)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def create_workbook(title):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Übung'
    ws.sheet_properties.tabColor = '1F4E79'
    ws.merge_cells('A1:F1')
    ws['A1'] = title
    ws['A1'].font = Font(name='Calibri', bold=True, size=14, color='1F4E79')
    ws['A1'].alignment = Alignment(horizontal='left')
    ws.merge_cells('A3:F3')
    ws['A3'] = 'Anleitung:'
    ws['A3'].font = Font(name='Calibri', bold=True, size=12)
    return wb, ws

def write_instructions(ws, instructions, start_row=4):
    for i, line in enumerate(instructions):
        row = start_row + i
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = line
        ws[f'A{row}'].font = INSTR_FONT
        ws[f'A{row}'].fill = INSTR_FILL
    return start_row + len(instructions) + 1

def write_table(ws, headers, data, start_row):
    for r in range(start_row, start_row + len(data) + 2):
        try: ws.unmerge_cells(f'A{r}:F{r}')
        except ValueError: pass
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=ci, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center'); cell.border = THIN_BORDER
    for ri, row_data in enumerate(data, start_row + 1):
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = DATA_FONT; cell.border = THIN_BORDER
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal='right')
    for ci in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(ci)].width = max(14, len(str(headers[ci-1]))+4)
    return start_row + len(data) + 1

def save(wb, name):
    wb.save(os.path.join(OUT_DIR, f'{name}.xlsx'))

# ============================================================
# MODUL 1: Erweiterte Formate
# ============================================================
def mod1_1():
    wb, ws = create_workbook('Übung 1.1 — Benutzerdefinierte Zahlenformate')
    headers = ['Produkt', 'Preis', 'Menge', 'Umsatz']
    data = [['Laptop', 1200, 5, None], ['Monitor', 350, 12, None], ['Tastatur', 80, 25, None]]
    write_table(ws, headers, data, 7)
    write_instructions(ws, ['1. Formatieren Sie Spalte B mit: #.##0,00 € (positiv); [Rot]-#.##0,00 € (negativ)',
        '2. Formatieren Sie Spalte D benutzerdefiniert: Werte > 5000 als 5,0 T anzeigen.',
        '3. Testen Sie mit negativen Werten und Null.'])
    save(wb, 'M1_1_Zahlenformate')

def mod1_2():
    wb, ws = create_workbook('Übung 1.2 — Bedingte Formatierung mit Formeln')
    headers = ['Auftrag', 'Betrag', 'Status', 'Fällig']
    data = [['A001', 8500, 'Offen', '15.08.2026'], ['A002', 12500, 'Offen', '20.09.2026'],
            ['A003', 5000, 'Erledigt', '01.06.2026'], ['A004', 18000, 'Offen', '05.08.2026']]
    write_table(ws, headers, data, 7)
    write_instructions(ws, ['1. Heben Sie Zeilen mit Betrag > 10000 gelb hervor (= $B2>10000).',
        '2. Markieren Sie "Offen"-Zeilen mit rotem Text (= $C2="Offen").',
        '3. Heben Sie Fälligkeitsdaten in den nächsten 7 Tagen orange hervor.'])
    save(wb, 'M1_2_Bedingte_Formatierung')

def mod1_3():
    wb, ws = create_workbook('Übung 1.3 — Datenüberprüfung')
    headers = ['Mitarbeiter', 'Abteilung', 'Gehalt', 'Eintritt']
    data = [['', '', '', '']]
    write_table(ws, headers, data, 7)
    write_instructions(ws, ['Richten Sie folgende Validierungen ein:',
        'Spalte B: Dropdown aus Bereich "Abteilungen" (IT, Vertrieb, HR, Finanzen).',
        'Spalte C: Ganze Zahl zwischen 30000 und 120000.',
        'Spalte D: Datum ab 01.01.2020. Fügen Sie Eingabemeldungen hinzu.'])
    save(wb, 'M1_3_Validierung')

def mod1_4():
    wb, ws = create_workbook('Übung 1.4 — Schutz einrichten')
    headers = ['Posten', 'Betrag', 'MwSt (19%)', 'Brutto']
    data = [['Büromaterial', 250, None, None], ['Reisekosten', 480, None, None]]
    write_table(ws, headers, data, 7)
    write_instructions(ws, ['1. Spalte C und D enthalten Formeln (=B2*0,19 und =B2+C2).',
        '2. Entsperren Sie Spalte A und B (Eingabezellen).',
        '3. Aktivieren Sie den Blattschutz und blenden Sie die Formeln aus.'])
    save(wb, 'M1_4_Schutz')

# ============================================================
# MODUL 2: Erweiterte Funktionen
# ============================================================
def mod2_1():
    wb, ws = create_workbook('Übung 2.1 — INDEX + VERGLEICH')
    headers = ['Produkt', 'Kategorie', 'Preis', 'Lager']
    data = [['Laptop', 'IT', 1200, 15], ['Monitor', 'IT', 350, 30], ['Drucker', 'IT', 280, 8],
            ['Schreibtisch', 'Möbel', 450, 12], ['Stuhl', 'Möbel', 220, 25]]
    write_table(ws, headers, data, 7)
    write_instructions(ws, ['1. Finden Sie den Preis von "Drucker" mit INDEX+VERGLEICH.',
        '2. Suchen Sie nach "Stuhl" → Kategorie soll ausgegeben werden (rechts→links!).',
        '3. Warum kann SVERWEIS Aufgabe 2 nicht lösen?'])
    save(wb, 'M2_1_INDEX_VERGLEICH')

def mod2_2():
    wb, ws = create_workbook('Übung 2.2 — Dynamische Bezüge')
    headers = ['Monat', 'Umsatz']
    data = [['Januar', 5000], ['Februar', 6200], ['März', 5800], ['April', 7100],
            ['Mai', 6900], ['Juni', 7500], ['Juli', 7200], ['August', 8000]]
    write_table(ws, headers, data, 7)
    write_instructions(ws, ['1. Erstellen Sie eine dynamische Summe mit BEREICH.VERSCHIEBEN.',
        '2. Formel für Durchschnitt der letzten 3 Monate mit dynamischem Bereich.',
        '3. Fügen Sie eine neue Zeile hinzu — passt sich die Summe an?'])
    save(wb, 'M2_2_BEREICH_VERSCHIEBEN')

def mod2_3():
    wb, ws = create_workbook('Übung 2.3 — Verschachtelte Funktionen')
    headers = ['Verkäufer', 'Umsatz', 'Provision']
    data = [['Anna', 8500, None], ['Tom', 25000, None], ['Lisa', 52000, None], ['Max', 12000, None]]
    write_table(ws, headers, data, 7)
    write_instructions(ws, ['1. Provision: <10000=5%, 10000-50000=8%, >50000=12% (verschachteltes WENN).',
        '2. Gleiche Logik mit WENNS (Excel 2019+) falls verfügbar.',
        '3. Fügen Sie WENNFEHLER hinzu, falls der Name nicht in einer Referenztabelle steht.'])
    save(wb, 'M2_3_Logik')

def mod2_4():
    wb, ws = create_workbook('Übung 2.4 — Finanzfunktionen')
    write_instructions(ws, ['Berechnen Sie in separaten Zellen:',
        '1. RMZ: Monatsrate für 250.000€ Kredit, 4,5% Zins, 30 Jahre.',
        '2. NBW: Investition 100.000€, 6 Jahre Rückflüsse à 25.000€, Zins 8%.',
        '3. ZW: Endkapital bei 200€/Monat, 3% p.a., 20 Jahre, 10.000€ Startkapital.'])
    save(wb, 'M2_4_Finanzfunktionen')

def mod2_5():
    wb, ws = create_workbook('Übung 2.5 — Matrixformeln')
    headers = ['Region', 'Produkt', 'Umsatz']
    data = [['Nord', 'A', 1200], ['Süd', 'B', 800], ['Nord', 'B', 2100],
            ['Ost', 'A', 950], ['Süd', 'A', 1600], ['Nord', 'C', 750]]
    write_table(ws, headers, data, 7)
    write_instructions(ws, ['1. Matrixformel: Summe aller Umsätze > 1000.',
        '2. Matrixformel: Größter Umsatz für Region "Nord".',
        '3. (Excel 365): =SORTIEREN(EINDEUTIG(A2:A7)) für eindeutige Regionen.'])
    save(wb, 'M2_5_Matrixformeln')

# ============================================================
# MODUL 3-10 exercises (abbreviated — core exercises per module)
# ============================================================

def mod2_6():
    wb, ws = create_workbook('Übung 2.6 — Datums- und Zeitfunktionen')
    headers = ['Mitarbeiter', 'Geburtsdatum', 'Eintritt', 'Alter', 'Betriebszugehörigkeit']
    data = [['Anna', '15.03.1985', '01.06.2018', None, None],
            ['Tom', '22.11.1990', '15.03.2020', None, None],
            ['Lisa', '08.07.1978', '01.01.2015', None, None]]
    write_table(ws, headers, data, 7)
    write_instructions(ws, ['1. Alter: =BRTEILJAHRE(B2;HEUTE())',
        '2. Betriebszugehörigkeit in Jahren: =BRTEILJAHRE(C2;HEUTE())',
        '3. Letzter Tag des aktuellen Monats: =MONATSENDE(HEUTE();0)',
        '4. Fälligkeitsdatum 30 Arbeitstage nach Eintritt: =ARBEITSTAG(C2;30)'])
    save(wb, 'M2_6_Datum_Zeit')

def mod3_1():
    wb, ws = create_workbook('Übung 3.1 — Namen definieren')
    headers = ['Produkt', 'Preis', 'MwSt-Satz']
    data = [['Laptop', 1200, None], ['Monitor', 350, None]]
    write_table(ws, headers, data, 7)
    write_instructions(ws, ['1. Definieren Sie den Namen "MwSt" für die Zelle mit 19%.',
        '2. Berechnen Sie Bruttopreise mit =B2*(1+MwSt).',
        '3. Definieren Sie "Preisliste" für A2:A3 und verwenden Sie den Namen in SVERWEIS.'])
    save(wb, 'M3_1_Namen')

def mod3_2():
    wb, ws = create_workbook('Übung 3.2 — 3D-Bezüge')
    for name, vals in [('Jan', 5000), ('Feb', 6200), ('Mär', 5800)]:
        s = wb.create_sheet(name)
        s['A1'], s['B1'] = 'Monat', 'Umsatz'
        s['A2'], s['B2'] = name, vals
    ws = wb['Übung']
    ws['A1'] = 'Jahresübersicht'
    write_instructions(ws, ['1. Auf Blatt "Jahresübersicht": =SUMME(Jan:Mär!B2).',
        '2. Fügen Sie ein neues Blatt zwischen Jan und Feb ein — wird es automatisch summiert?',
        '3. Testen Sie MITTELWERT und MAX mit 3D-Bezügen.'])
    save(wb, 'M3_2_3D_Bezuege')

def mod3_3():
    wb, ws = create_workbook('Übung 3.3 — Verknüpfungen')
    write_instructions(ws, ['1. Erstellen Sie eine neue Arbeitsmappe "Quelle.xlsx" mit Daten.',
        '2. Verknüpfen Sie in dieser Übung auf die externe Zelle.',
        '3. Aktualisieren Sie die Verknüpfung und testen Sie "Verknüpfung lösen".'])
    save(wb, 'M3_3_Verknuepfungen')


def mod3_4():
    wb, ws = create_workbook('Übung 3.4 — Daten konsolidieren')
    for name, vals in [('Q1', [('A',1000),('B',1500),('C',800)]),
                        ('Q2', [('C',900),('A',1200),('B',1400)]),
                        ('Q3', [('B',1600),('A',1100),('C',850)])]:
        s = wb.create_sheet(name)
        s['A1'], s['B1'] = 'Produkt', 'Umsatz'
        for ri, (prod, val) in enumerate(vals, 2):
            s.cell(row=ri, column=1, value=prod); s.cell(row=ri, column=2, value=val)
    ws = wb['Übung']
    write_instructions(ws, ['1. Wechseln Sie zum Blatt "Jahresübersicht" (neu erstellen).',
        '2. Daten → Konsolidieren: Summe über Q1, Q2, Q3.',
        '3. Aktivieren Sie "Verknüpfung mit Quelldaten" und ändern Sie einen Wert in Q1.',
        '4. Beobachten Sie: Aktualisiert sich die Konsolidierung automatisch?'])
    save(wb, 'M3_4_Konsolidierung')

def mod4_1():
    wb, ws = create_workbook('Übung 4.1 — Spezialfilter')
    headers = ['Region', 'Produkt', 'Umsatz', 'Datum']
    data = [['Nord', 'A', 12000, '15.01.2026'], ['Süd', 'B', 8000, '20.01.2026'],
            ['Nord', 'C', 15000, '01.02.2026'], ['Ost', 'A', 5000, '10.02.2026'],
            ['Süd', 'A', 11000, '25.02.2026'], ['Nord', 'B', 9000, '05.03.2026']]
    write_table(ws, headers, data, 7)
    write_instructions(ws, ['1. Spezialfilter: Region=Nord UND Umsatz>10000.',
        '2. Erweiterte Kriterien: (Nord UND >10000) ODER (Süd UND >10000).',
        '3. Extrahieren Sie das Ergebnis in einen neuen Bereich.'])
    save(wb, 'M4_1_Spezialfilter')

def mod4_2():
    wb, ws = create_workbook('Übung 4.2 — Datenbankfunktionen')
    headers = ['Kunde', 'Region', 'Umsatz', 'Alter']
    data = [['Müller', 'West', 5000, 35], ['Schmidt', 'Ost', 8000, 42],
            ['Weber', 'West', 3000, 28], ['Fischer', 'Nord', 12000, 55]]
    write_table(ws, headers, data, 7)
    write_instructions(ws, ['1. DBSUMME: Gesamtumsatz Region "West".',
        '2. DBMITTELWERT: Durchschnittsalter Region "West".',
        '3. DBAUSZUG: Name des Kunden mit Umsatz > 10000.'])
    save(wb, 'M4_2_Datenbankfunktionen')

def mod4_3():
    wb, ws = create_workbook('Übung 4.3 — Teilergebnisse')
    headers = ['Region', 'Produkt', 'Umsatz']
    data = [['Nord', 'A', 1200], ['Nord', 'B', 800], ['Süd', 'A', 1500],
            ['Süd', 'B', 900], ['Ost', 'A', 1100]]
    write_table(ws, headers, data, 7)
    write_instructions(ws, ['1. Sortieren Sie nach Region, dann Produkt.',
        '2. Fügen Sie Teilergebnisse ein: Summe Umsatz pro Region.',
        '3. Fügen Sie eine zweite Ebene: Anzahl pro Produkt.'])
    save(wb, 'M4_3_Teilergebnisse')


def mod4_4():
    wb, ws = create_workbook('Übung 4.4 — Excel-Tabellen')
    headers = ['Produkt', 'Menge', 'Preis', 'Summe']
    data = [['Laptop', 5, 1200, None], ['Monitor', 12, 350, None], ['Tastatur', 25, 80, None]]
    write_table(ws, headers, data, 7)
    write_instructions(ws, ['1. Strg+T: Bereich als Tabelle formatieren.',
        '2. Spalte Summe mit strukturiertem Verweis: =[@Menge]*[@Preis].',
        '3. Ergebniszeile aktivieren: Summe der Menge anzeigen.',
        '4. Neue Zeile einfügen: Werden Formatierung und Formeln automatisch übernommen?'])
    save(wb, 'M4_4_Tabellen')

def mod5_1():
    wb, ws = create_workbook('Übung 5.1 — Pivot-Tabelle')
    headers = ['Region', 'Produkt', 'Quartal', 'Umsatz']
    data = [['Nord', 'A', 'Q1', 5000], ['Nord', 'B', 'Q1', 3000], ['Süd', 'A', 'Q1', 4000],
            ['Nord', 'A', 'Q2', 6000], ['Süd', 'B', 'Q2', 3500]]
    write_table(ws, headers, data, 7)
    write_instructions(ws, ['1. Erstellen Sie eine Pivot-Tabelle: Region+Produkt in Zeilen, Quartal in Spalten.',
        '2. Werte: Summe Umsatz. Ändern Sie auf Mittelwert.',
        '3. Zeigen Sie Werte als % des Gesamtergebnisses an.'])
    save(wb, 'M5_1_Pivot')

def mod5_2():
    wb, ws = create_workbook('Übung 5.2 — Pivot-Anpassung')
    headers = ['Datum', 'Region', 'Produkt', 'Umsatz', 'Kosten']
    data = [['15.01.2026', 'Nord', 'A', 5000, 3000], ['20.02.2026', 'Süd', 'B', 4000, 2500],
            ['10.03.2026', 'Nord', 'A', 6000, 3500]]
    write_table(ws, headers, data, 7)
    write_instructions(ws, ['1. Gruppieren Sie das Datumsfeld nach Monaten und Quartalen.',
        '2. Erstellen Sie ein berechnetes Feld Marge = (Umsatz-Kosten)/Umsatz.',
        '3. Formatieren Sie Marge als Prozent.'])
    save(wb, 'M5_2_Pivot_Anpassung')

def mod5_3():
    wb, ws = create_workbook('Übung 5.3 — Slicer')
    headers = ['Region', 'Produkt', 'Kategorie', 'Umsatz']
    data = [['Nord', 'Laptop', 'IT', 5000], ['Süd', 'Monitor', 'IT', 3000], ['Nord', 'Stuhl', 'Möbel', 2000]]
    write_table(ws, headers, data, 7)
    write_instructions(ws, ['1. Erstellen Sie eine Pivot-Tabelle und fügen Sie Slicer für Region und Kategorie ein.',
        '2. Erstellen Sie eine zweite Pivot-Tabelle (Anzahl).',
        '3. Verbinden Sie beide Pivot-Tabellen mit denselben Slicern.'])
    save(wb, 'M5_3_Slicer')

def mod5_4():
    wb, ws = create_workbook('Übung 5.4 — Pivot-Chart')
    headers = ['Monat', 'Umsatz', 'Region']
    data = [['Jan', 5000, 'Nord'], ['Feb', 6200, 'Nord'], ['Mär', 5800, 'Nord']]
    write_table(ws, headers, data, 7)
    write_instructions(ws, ['1. Erstellen Sie aus der Pivot-Tabelle ein Pivot-Chart (Säulen).',
        '2. Fügen Sie einen Slicer für Region hinzu.',
        '3. Wechseln Sie die Region im Slicer und beobachten Sie das Diagramm.'])
    save(wb, 'M5_4_PivotChart')

def mod6_1():
    wb, ws = create_workbook('Übung 6.1 — Zielwertsuche')
    write_instructions(ws, ['Szenario: Sie verkaufen 500 Einheiten zu je 50 €. Fixkosten: 10.000 €.',
        'Zelle A1: Menge (500), B1: Preis (50), C1: =A1*B1 (Umsatz), D1: =C1-10000 (Gewinn).',
        '1. Zielwertsuche: Welcher Preis ist nötig für 50.000 € Gewinn?',
        '2. Welche Menge bei 50 € Stückpreis für 50.000 € Gewinn?'])
    save(wb, 'M6_1_Zielwertsuche')

def mod6_2():
    wb, ws = create_workbook('Übung 6.2 — Szenarien')
    write_instructions(ws, ['Modell: Gewinn = Menge * Preis - Fixkosten.',
        'Erstellen Sie drei Szenarien mit dem Szenario-Manager:',
        '1. Optimistisch: Menge=1000, Preis=60, Fixkosten=8000',
        '2. Neutral: Menge=500, Preis=50, Fixkosten=10000',
        '3. Pessimistisch: Menge=300, Preis=45, Fixkosten=12000',
        'Erstellen Sie einen Szenario-Zusammenfassungsbericht.'])
    save(wb, 'M6_2_Szenarien')

def mod6_3():
    wb, ws = create_workbook('Übung 6.3 — Solver')
    headers = ['Produkt', 'Gewinn/Stück', 'Max. Menge', 'Produktion']
    data = [['A', 50, 100, None], ['B', 80, 80, None], ['C', 65, 120, None]]
    write_table(ws, headers, data, 7)
    write_instructions(ws, ['Gesamtkapazität: 200 Einheiten. Ziel: Gewinn maximieren.',
        '1. Aktivieren Sie den Solver (Add-Ins).',
        '2. Zielzelle: Gesamtgewinn, Max. Veränderbare Zellen: Produktion.',
        '3. Nebenbedingungen: Produktion <= Max. Menge, Summe Produktion <= 200, >= 0.'])
    save(wb, 'M6_3_Solver')

def mod6_4():
    wb, ws = create_workbook('Übung 6.4 — Sparklines')
    headers = ['Monat', 'Umsatz', 'Kosten', 'Gewinn']
    data = [['Jan', 5000, 3200, None], ['Feb', 6200, 3800, None], ['Mär', 5800, 3500, None],
            ['Apr', 7100, 4000, None], ['Mai', 6900, 4100, None], ['Jun', 7500, 4200, None]]
    write_table(ws, headers, data, 7)
    write_instructions(ws, ['1. Fügen Sie Liniensparklines für die Umsatzspalte ein.',
        '2. Erstellen Sie ein Liniendiagramm und fügen Sie eine Trendlinie mit R² hinzu.',
        '3. Markieren Sie Höchst- und Tiefstwerte in den Sparklines.'])
    save(wb, 'M6_4_Sparklines')

def mod7_1():
    wb, ws = create_workbook('Übung 7.1 — Verbunddiagramm')
    headers = ['Monat', 'Umsatz', 'Wachstum %']
    data = [['Jan', 50000, None], ['Feb', 52000, None], ['Mär', 55000, None],
            ['Apr', 53000, None], ['Mai', 58000, None], ['Jun', 60000, None]]
    write_table(ws, headers, data, 7)
    write_instructions(ws, ['1. Berechnen Sie die Wachstumsrate in Spalte C: =(B3-B2)/B2.',
        '2. Erstellen Sie ein Kombidiagramm: Umsatz als Säulen, Wachstum als Linie.',
        '3. Fügen Sie eine Sekundärachse für die Wachstumsrate ein.'])
    save(wb, 'M7_1_Verbunddiagramm')

def mod7_2():
    wb, ws = create_workbook('Übung 7.2 — Wasserfalldiagramm')
    headers = ['Position', 'Betrag']
    data = [['Anfangsbestand', 100000], ['Einnahmen Q1', 25000], ['Ausgaben Q1', -15000],
            ['Einnahmen Q2', 30000], ['Ausgaben Q2', -20000], ['Endbestand', None]]
    write_table(ws, headers, data, 7)
    write_instructions(ws, ['1. Erstellen Sie ein Wasserfalldiagramm aus diesen Daten.',
        '2. Formatieren Sie: Erhöhungen grün, Verminderungen rot.',
        '3. Setzen Sie den Endbestand als Gesamtwert (blau).'])
    save(wb, 'M7_2_Wasserfall')

def mod7_3():
    wb, ws = create_workbook('Übung 7.3 — Dashboard')
    headers = ['Monat', 'Region', 'Produkt', 'Umsatz']
    data = [['Jan', 'Nord', 'A', 5000], ['Jan', 'Süd', 'B', 3000], ['Feb', 'Nord', 'A', 6000],
            ['Feb', 'Süd', 'B', 3500], ['Mär', 'Nord', 'A', 5500], ['Mär', 'Süd', 'B', 4000]]
    write_table(ws, headers, data, 7)
    write_instructions(ws, ['Erstellen Sie auf einem NEUEN Blatt ein Dashboard mit:',
        '1. Liniendiagramm: Umsatzverlauf über 6 Monate.',
        '2. Säulendiagramm: Umsatz nach Region.',
        '3. Slicer für Region und Produkt.',
        '4. Ordnen Sie die Elemente übersichtlich an.'])
    save(wb, 'M7_3_Dashboard')

def mod8_1():
    wb, ws = create_workbook('Übung 8.1 — Makro aufzeichnen')
    headers = ['Monat', 'Umsatz', 'Kosten']
    data = [['Januar', 5000, 3200], ['Februar', 6200, 3800], ['März', 5800, 3500]]
    write_table(ws, headers, data, 7)
    write_instructions(ws, ['1. Aktivieren Sie die Entwicklertools und speichern Sie als .xlsm.',
        '2. Zeichnen Sie ein Makro "FormatBericht" auf: Überschrift fett+blau, Rahmen, Summenzeile.',
        '3. Führen Sie das Makro auf einem zweiten Blatt aus.'])
    save(wb, 'M8_1_Makro_Aufzeichnen')

def mod8_2():
    wb, ws = create_workbook('Übung 8.2 — Makro zuweisen')
    write_instructions(ws, ['1. Weisen Sie das Makro aus Übung 8.1 einer Schaltfläche zu.',
        '2. Richten Sie Strg+Umschalt+F als Tastenkombination ein.',
        '3. Testen Sie beide Ausführungsmethoden auf neuen Daten.'])
    save(wb, 'M8_2_Makro_Zuweisen')

def mod8_3():
    wb, ws = create_workbook('Übung 8.3 — VBA-Editor')
    write_instructions(ws, ['1. Öffnen Sie den VBA-Editor mit Alt+F11.',
        '2. Finden Sie das Makro "FormatBericht" und ändern Sie die Farbe.',
        '3. Fügen Sie eine MsgBox "Formatierung abgeschlossen!" am Ende des Makros ein.'])
    save(wb, 'M8_3_VBA_Editor')

def mod9_1():
    wb, ws = create_workbook('Übung 9.1 — VBA Variablen')
    write_instructions(ws, ['1. Schreiben Sie ein Makro, das Werte aus B2 und C2 liest und das Produkt in D2 schreibt.',
        '2. Deklarieren Sie Variablen mit Dim (Double für Zahlen).',
        '3. Geben Sie das Ergebnis zusätzlich mit MsgBox aus.'])
    save(wb, 'M9_1_VBA_Variablen')

def mod9_2():
    wb, ws = create_workbook('Übung 9.2 — VBA Kontrollstrukturen')
    write_instructions(ws, ['1. For-Schleife: Zahlen 1-10 in A1-A10 schreiben.',
        '2. Erweitern: If-Bedingung — Zahlen > 5 werden fett formatiert.',
        '3. For-Each-Schleife: Alle Zellen mit Wert > 1000 in Spalte B gelb markieren.'])
    save(wb, 'M9_2_VBA_Kontrollstrukturen')

def mod9_3():
    wb, ws = create_workbook('Übung 9.3 — VBA Ereignisse')
    write_instructions(ws, ['1. Worksheet_Change: MsgBox bei Wert > 10000 in Spalte B.',
        '2. Workbook_Open: Heutiges Datum in A1 beim Öffnen eintragen.',
        '3. Testen Sie beide Ereignisse durch Ändern von Werten und Neustart.'])
    save(wb, 'M9_3_VBA_Ereignisse')

def mod9_4():
    wb, ws = create_workbook('Übung 9.4 — VBA UDF')
    write_instructions(ws, ['1. Schreiben Sie eine UDF "Bonus(Umsatz)": 5% ab 10000, sonst 0%.',
        '2. Verwenden Sie =Bonus(B2) in einer Zelle.',
        '3. Erstellen Sie "Kategorie(Alter)": <30="Junior", 30-50="Senior", >50="Experte".'])
    save(wb, 'M9_4_VBA_UDF')

def mod10_1():
    wb, ws = create_workbook('Übung 10.1 — Vorlagen')
    write_instructions(ws, ['1. Erstellen Sie eine Rechnungsvorlage mit Firmenlogo-Bereich und automatischer Nummer.',
        '2. Fügen Sie geschützte Formelzellen für MwSt-Berechnung hinzu.',
        '3. Speichern Sie als .xltx und öffnen Sie die Vorlage erneut.'])
    save(wb, 'M10_1_Vorlagen')

def mod10_2():
    wb, ws = create_workbook('Übung 10.2 — Zusammenarbeit')
    headers = ['Posten', 'Betrag']
    data = [['Beratung', 1500], ['Schulung', 800]]
    write_table(ws, headers, data, 7)
    write_instructions(ws, ['1. Fügen Sie einen Kommentar zu Zelle B2 ein.',
        '2. Exportieren Sie das Blatt als PDF.',
        '3. Schützen Sie die Überschriften, lassen Sie Eingabezellen bearbeitbar.'])
    save(wb, 'M10_2_Zusammenarbeit')

def mod10_3():
    wb, ws = create_workbook('Übung 10.3 — Tastenkombinationen')
    headers = ['Monat', 'Umsatz', 'Kosten']
    data = [['Jan', 5000, 3200], ['Feb', 6200, 3800], ['Mär', 5800, 3500]]
    write_table(ws, headers, data, 7)
    write_instructions(ws, ['Arbeiten Sie NUR mit Tastatur:',
        '1. Strg+Shift+L: Filter aktivieren. Alt+=: Summe einfügen.',
        '2. Strg+1: Zellen formatieren. F4: Letzte Aktion wiederholen.',
        '3. Strg+[: Vorgängerzellen anzeigen. Strg+]: Nachfolger anzeigen.'])
    save(wb, 'M10_3_Tastenkombinationen')

if __name__ == '__main__':
    print('Modul 1: Erweiterte Formate...')
    mod1_1(); mod1_2(); mod1_3(); mod1_4()
    print('Modul 2: Erweiterte Funktionen...')
    mod2_1(); mod2_2(); mod2_3(); mod2_4(); mod2_5(); mod2_6()
    print('Modul 3: Referenzen & Namen...')
    mod3_1(); mod3_2(); mod3_3(); mod3_4()
    print('Modul 4: Datenbanken...')
    mod4_1(); mod4_2(); mod4_3(); mod4_4()
    print('Modul 5: Pivot-Tabellen...')
    mod5_1(); mod5_2(); mod5_3(); mod5_4()
    print('Modul 6: Datenanalyse...')
    mod6_1(); mod6_2(); mod6_3(); mod6_4()
    print('Modul 7: Diagramme & Dashboards...')
    mod7_1(); mod7_2(); mod7_3()
    print('Modul 8: Makros...')
    mod8_1(); mod8_2(); mod8_3()
    print('Modul 9: VBA...')
    mod9_1(); mod9_2(); mod9_3(); mod9_4()
    print('Modul 10: Zusammenarbeit...')
    mod10_1(); mod10_2(); mod10_3()
    print(f'\nFertig! Dateien in {OUT_DIR}')
