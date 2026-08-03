#!/usr/bin/env python3
"""Generate XLSX exercise files for Excel-lenz beginner course."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT_DIR = os.path.join(os.path.dirname(__file__), "../exercises/anfaenger")
os.makedirs(OUT_DIR, exist_ok=True)

HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
INSTR_FILL = PatternFill('solid', fgColor='F2F2F2')
INSTR_FONT = Font(name='Calibri', italic=True, color='333333', size=10)
DATA_FONT = Font(name='Calibri', size=11)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def create_workbook(title):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Übung'
    ws.sheet_properties.tabColor = '1F4E79'

    # Title row
    ws.merge_cells('A1:F1')
    ws['A1'] = title
    ws['A1'].font = Font(name='Calibri', bold=True, size=14, color='1F4E79')
    ws['A1'].alignment = Alignment(horizontal='left')

    # Instructions area (row 3)
    ws.merge_cells('A3:F3')
    ws['A3'] = '📝 Anleitung:'
    ws['A3'].font = Font(name='Calibri', bold=True, size=12)
    return wb, ws


def write_instructions(ws, instructions, start_row=4):
    for i, line in enumerate(instructions):
        row = start_row + i
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = line
        ws[f'A{row}'].font = INSTR_FONT
        ws[f'A{row}'].fill = INSTR_FILL
    return start_row + len(instructions) + 1


def write_table(ws, headers, data, start_row):
    # Unmerge cells in the table range to avoid MergedCell errors
    for r in range(start_row, start_row + len(data) + 2):
        try:
            ws.unmerge_cells(f'A{r}:F{r}')
        except ValueError:
            pass
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    for ri, row_data in enumerate(data, start_row + 1):
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal='right')

    # Auto-width
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = max(14, len(str(headers[ci - 1])) + 4)

    return start_row + len(data) + 1


def add_result_row(ws, label, formula, row, col=1):
    ws.cell(row=row, column=col, value=label).font = Font(name='Calibri', bold=True, size=11)
    ws.cell(row=row, column=col + 1, value=formula).font = Font(name='Consolas', size=11, color='1F4E79')
    return row + 1


# ============================================================
# MODUL 1
# ============================================================

def mod1_1():
    wb, ws = create_workbook('Übung 1.1 — Erste Schritte')
    write_instructions(ws, [
        '1. Öffnen Sie Excel und erstellen Sie eine neue Arbeitsmappe.',
        '2. Identifizieren Sie: Registerkarten, Namensfeld, Bearbeitungsleiste, Statusleiste.',
        '3. Speichern Sie die Arbeitsmappe als Meine_Erste_Mappe.xlsx.',
        '✅ Diese Übung dient der Orientierung — es gibt keine vordefinierten Daten.'
    ])
    save(wb, 'M1_1_Erste_Schritte')


def mod1_2():
    wb, ws = create_workbook('Übung 1.2 — Die Oberfläche erkunden')
    write_instructions(ws, [
        '1. Fügen Sie der Symbolleiste für den Schnellzugriff „Neu", „Öffnen" und „Schnelldruck" hinzu.',
        '2. Finden Sie die Gruppe „Schriftart" auf der Registerkarte „Start".',
        '3. Ändern Sie den Zoom auf 150% und zurück auf 100%.',
        '✅ Orientierungsübung ohne vordefinierte Daten.'
    ])
    save(wb, 'M1_2_Oberflaeche')


def mod1_3():
    wb, ws = create_workbook('Übung 1.3 — Navigation üben')
    write_instructions(ws, [
        '1. Erstellen Sie drei Tabellenblätter: Januar, Februar, März.',
        '2. Navigieren Sie mit Strg+Pos1 zur Zelle A1.',
        '3. Navigieren Sie mit Strg+Ende zur letzten verwendeten Zelle.',
        '4. Navigieren Sie mit Strg+↓ zum Ende des Datenbereichs.',
        '✅ Navigationsübung ohne vordefinierte Daten.'
    ])
    save(wb, 'M1_3_Navigation')


def mod1_4():
    wb, ws = create_workbook('Übung 1.4 — Dateien verwalten')
    write_instructions(ws, [
        '1. Speichern Sie diese Arbeitsmappe als Inventar_2026.xlsx im Ordner Dokumente.',
        '2. Exportieren Sie die Arbeitsmappe als PDF.',
        '3. Schließen und öffnen Sie die Arbeitsmappe erneut.',
        '✅ Dateiverwaltungsübung ohne vordefinierte Daten.'
    ])
    save(wb, 'M1_4_Dateiverwaltung')


# ============================================================
# MODUL 2
# ============================================================

def mod2_1():
    wb, ws = create_workbook('Übung 2.1 — Datentypen erkennen')
    data = [
        ['Excel-Schulung'],
        [1500],
        ['15.03.2026'],
        ['25,50 €'],
        ['10:30'],
    ]
    headers = ['A — Einträge']
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        'Geben Sie folgende Daten in Spalte A ein und beobachten Sie die Ausrichtung:',
        'Excel-Schulung | 1500 | 15.03.2026 | 25,50 € | 10:30',
        'Welche Einträge sind linksbündig (Text), welche rechtsbündig (Zahlen/Datum)?'
    ])
    save(wb, 'M2_1_Datentypen')


def mod2_2():
    wb, ws = create_workbook('Übung 2.2 — Daten bearbeiten')
    headers = ['Produkt', 'Menge', 'Preis']
    data = [
        ['Apfel', 10, 0.50],
        ['Birne', 15, 0.60],
        ['Kirsche', 20, 0.80],
    ]
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        'Gegeben ist die Tabelle unten. Führen Sie folgende Änderungen durch:',
        '1. Ändern Sie den Preis von Kirsche auf 0,75 €.',
        '2. Fügen Sie eine neue Zeile für Banane ein (Menge: 12, Preis: 0,45 €).',
        '3. Löschen Sie die Zeile Birne, dann machen Sie es rückgängig (Strg+Z).'
    ])
    save(wb, 'M2_2_Bearbeiten')


def mod2_3():
    wb, ws = create_workbook('Übung 2.3 — AutoAusfüllen')
    write_instructions(ws, [
        'Führen Sie folgende AutoAusfüllen-Aufgaben durch:',
        '1. In A1: „Januar" eingeben und bis A12 ziehen → Monate Januar–Dezember.',
        '2. In B1: „Montag" eingeben und nach rechts bis H1 ziehen → Wochentage.',
        '3. In C1: 5, in C2: 10, beide markieren und bis C10 ziehen → 5er-Schritte.',
        '4. In D1: „1. Januar 2026" und bis D31 ziehen → alle Januartage.'
    ])
    save(wb, 'M2_3_AutoAusfuellen')


def mod2_4():
    wb, ws = create_workbook('Übung 2.4 — Kopieren und Einfügen')
    headers = ['Name', 'Alter', 'Stadt']
    data = [
        ['Anna', 28, 'Berlin'],
        ['Tom', 35, 'München'],
    ]
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        '1. Kopieren Sie die Tabelle nach E1 mit Strg+C / Strg+V.',
        '2. Kopieren Sie die Tabelle nach A5 mit „Inhalte einfügen → Transponieren".',
        '3. Kopieren Sie nur die Werte (ohne Formatierung) nach A10.'
    ])
    save(wb, 'M2_4_Kopieren')


# ============================================================
# MODUL 3
# ============================================================

def mod3_1():
    wb, ws = create_workbook('Übung 3.1 — Grundformatierung')
    headers = ['Monat', 'Einnahmen', 'Ausgaben', 'Gewinn']
    data = [
        ['Januar', 5000, 3200, None],
        ['Februar', 4800, 3100, None],
        ['März', 5200, 3300, None],
    ]
    row = write_table(ws, headers, data, 7)
    write_instructions(ws, [
        '1. Kopfzeile: Fett, zentriert, blaue Füllfarbe, weiße Schrift.',
        '2. Außenrahmen um die gesamte Tabelle.',
        '3. Alle Zahlen zentrieren.',
        '4. Titel „Monatsübersicht 2026" über die Tabellenbreite zentrieren (Zellen verbinden).'
    ])
    save(wb, 'M3_1_Grundformatierung')


def mod3_2():
    wb, ws = create_workbook('Übung 3.2 — Zahlen formatieren')
    headers = ['Monat', 'Einnahmen', 'Ausgaben', 'Gewinn']
    data = [
        ['Januar', 5000, 3200, None],
        ['Februar', 4800, 3100, None],
        ['März', 5200, 3300, None],
    ]
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        '1. Formatieren Sie Einnahmen & Ausgaben als Währung (€) mit Tausendertrennzeichen.',
        '2. Formatieren Sie die Gewinn-Spalte als Währung.',
        '3. Neue Zeile: Prozentsatz Gewinn/Einnahmen, formatiert als % mit 1 Dezimalstelle.'
    ])
    save(wb, 'M3_2_Zahlenformat')


def mod3_3():
    wb, ws = create_workbook('Übung 3.3 — Layout anpassen')
    headers = ['Monat', 'Einnahmen', 'Ausgaben', 'Gewinn']
    data = [
        ['Januar', 5000, 3200, None],
        ['Februar', 4800, 3100, None],
        ['März', 5200, 3300, None],
    ]
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        '1. Spaltenbreite automatisch anpassen (Doppelklick zwischen Spaltenköpfen).',
        '2. Spalte C (Ausgaben) ausblenden und wieder einblenden.',
        '3. Zeilenhöhe der Kopfzeile auf 30 Pixel setzen.'
    ])
    save(wb, 'M3_3_Layout')


def mod3_4():
    wb, ws = create_workbook('Übung 3.4 — Bedingte Formatierung')
    headers = ['Schüler', 'Note']
    data = [
        ['Anna', 1.7],
        ['Tom', 3.3],
        ['Lisa', 2.0],
        ['Max', 4.0],
        ['Sara', 1.3],
    ]
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        '1. Heben Sie alle Noten unter 2,0 grün hervor (bedingte Formatierung).',
        '2. Heben Sie alle Noten über 3,5 rot hervor.',
        '3. Fügen Sie grüne/gelbe/rote Farbskalen hinzu.',
        '4. Fügen Sie Ampelsymbole hinzu (Symbolsätze).'
    ])
    save(wb, 'M3_4_Bedingte_Formatierung')


# ============================================================
# MODUL 4
# ============================================================

def mod4_1():
    wb, ws = create_workbook('Übung 4.1 — Erste Formeln')
    write_instructions(ws, [
        'Geben Sie in A1 = 100 und B1 = 25 ein.',
        'Berechnen Sie in Spalte C folgende Formeln:',
        'C1: =A1+B1   (Addition → 125)',
        'C2: =A1-B1   (Subtraktion → 75)',
        'C3: =A1*B1   (Multiplikation → 2500)',
        'C4: =A1/B1   (Division → 4)',
        'C5: =A1^2    (Potenz → 10000)',
        'C6: =(A1+B1)*2 (Klammerrechnung → 250)'
    ], 3)
    ws['A12'] = 100
    ws['B12'] = 25
    ws['A12'].font = DATA_FONT
    ws['B12'].font = DATA_FONT
    save(wb, 'M4_1_Erste_Formeln')


def mod4_2():
    wb, ws = create_workbook('Übung 4.2 — Zellbezüge')
    headers = ['Produkt', 'Preis', 'Menge', 'Umsatz', 'MwSt-Satz', 'Bruttopreis']
    data = [
        ['A', 10, 5, None, 0.19, None],
        ['B', 20, 3, None, None, None],
        ['C', 15, 8, None, None, None],
    ]
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        'Der MwSt-Satz (19%) steht in Zelle E2.',
        '1. Berechnen Sie Umsatz in D2: =B2*C2 und ziehen Sie nach unten.',
        '2. Berechnen Sie Bruttopreis in F2: =D2*(1+$E$2) und ziehen Sie nach unten.',
        '3. Warum braucht E2 einen absoluten Bezug ($E$2)?'
    ])
    save(wb, 'M4_2_Zellbezuege')


def mod4_3_namen():
    wb, ws = create_workbook('Übung 4.3 — Namen definieren')
    headers = ['Produkt', 'Preis', 'MwSt-Satz']
    data = [
        ['Laptop', 1200, None],
        ['Monitor', 350, None],
        ['Tastatur', 80, None],
        ['Maus', 45, None],
        ['Drucker', 280, None],
    ]
    row = write_table(ws, headers, data, 7)
    # MwSt cell
    ws.cell(row=row, column=2, value='19%').font = Font(name='Calibri', bold=True, size=11)
    ws.cell(row=row, column=1, value='MwSt-Satz:').font = Font(name='Calibri', bold=True, size=11)
    row += 2
    write_instructions(ws, [
        '1. Definieren Sie für die Zelle mit 19% den Namen „MwSt_Satz" über das Namensfeld.',
        '2. Berechnen Sie in Spalte C den Bruttopreis: =B8*(1+MwSt_Satz)',
        '3. Definieren Sie den Namen „Preisliste" für den gesamten Datenbereich A7:A11.',
        '4. Prüfen Sie im Namens-Manager (Formeln → Namens-Manager) alle definierten Namen.',
    ], row)
    save(wb, 'M4_3_Namen')


def mod4_3():
    wb, ws = create_workbook('Übung 4.3 — Statistische Funktionen')
    headers = ['Produkt', 'Verkäufe']
    data = [
        ['Produkt A', 120],
        ['Produkt B', 85],
        ['Produkt C', 340],
        ['Produkt D', 67],
        ['Produkt E', 230],
        ['Produkt F', 450],
        ['Produkt G', 178],
        ['Produkt H', 500],
        ['Produkt I', 92],
        ['Produkt J', 310],
    ]
    row = write_table(ws, headers, data, 7) + 1
    add_result_row(ws, 'Gesamtsumme:', '=SUMME(B8:B17)', row)
    add_result_row(ws, 'Durchschnitt:', '=MITTELWERT(B8:B17)', row + 1)
    add_result_row(ws, 'Maximum:', '=MAX(B8:B17)', row + 2)
    add_result_row(ws, 'Minimum:', '=MIN(B8:B17)', row + 3)
    add_result_row(ws, 'Anzahl:', '=ANZAHL(B8:B17)', row + 4)
    save(wb, 'M4_4_Statistik')


def mod4_4():
    wb, ws = create_workbook('Übung 4.4 — WENN-Funktion')
    headers = ['Name', 'Punkte', 'Bestanden?']
    data = [
        ['Anna', 85, None],
        ['Tom', 42, None],
        ['Lisa', 67, None],
        ['Max', 91, None],
        ['Sara', 55, None],
    ]
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        '1. In C2: =WENN(B2>=60;"Ja";"Nein")',
        '2. Erweiterte Formel in C2:',
        '   =WENN(B2>=80;"Mit Auszeichnung";WENN(B2>=60;"Bestanden";"Nicht bestanden"))',
        '   Ziehen Sie die Formel nach unten.'
    ])
    save(wb, 'M4_5_WENN')


# ============================================================
# MODUL 5
# ============================================================

def mod5_1():
    wb, ws = create_workbook('Übung 5.1 — Datenvalidierung')
    headers = ['Produkt', 'Menge', 'Lieferdatum', 'Kunden-ID']
    data = [['', '', '', '']]
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        'Richten Sie folgende Datenvalidierungen ein:',
        'Spalte A (Produkt): Dropdown-Liste: Apfel, Birne, Banane, Orange',
        'Spalte B (Menge): Ganze Zahl zwischen 1 und 100',
        'Spalte C (Lieferdatum): Datum ab heute',
        'Spalte D (Kunden-ID): Textlänge = genau 5 Zeichen',
        'Fügen Sie Eingabemeldungen und Fehlerwarnungen hinzu.'
    ])
    save(wb, 'M5_1_Validierung')


def mod5_2():
    wb, ws = create_workbook('Übung 5.2 — Daten bereinigen')
    headers = ['Name', 'Stadt', 'Alter']
    data = [
        ['Max Mustermann', 'Berlin', 45],
        ['Anna Schmidt', 'München', 32],
        ['Max Mustermann', 'Berlin', 45],
        ['Tom Meyer', 'Hamburg', 28],
        ['Anna Schmidt', 'München', 32],
        ['Lisa Weber', 'Köln', 51],
    ]
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        '1. Die Daten wurden als Rohdaten importiert (siehe Tabelle).',
        '2. Entfernen Sie Duplikate (Daten → Duplikate entfernen).',
        '3. Extrahieren Sie Initialen aus Spalte A mit Flash Fill (Strg+E) — z.B. „MM".',
        '4. Ersetzen Sie „München" durch „Munich" mit Suchen & Ersetzen (Strg+H).'
    ])
    save(wb, 'M5_2_Bereinigen')


def mod5_3():
    wb, ws = create_workbook('Übung 5.3 — Konsolidierung')

    # Create Q1, Q2, Q3, Jahresübersicht as new sheets (not reusing ws)
    for qi, (name, vals) in enumerate([
        ('Q1', [('A', 1000), ('B', 1500), ('C', 800)]),
        ('Q2', [('A', 1200), ('B', 1400), ('C', 900)]),
        ('Q3', [('A', 1100), ('B', 1600), ('C', 850)]),
    ]):
        s = wb.create_sheet(name)
        s['A1'] = 'Produkt'
        s['B1'] = 'Umsatz'
        s['A1'].font = HEADER_FONT
        s['B1'].font = HEADER_FONT
        s['A1'].fill = s['B1'].fill = HEADER_FILL
        for ri, (prod, val) in enumerate(vals, 2):
            s.cell(row=ri, column=1, value=prod).font = DATA_FONT
            s.cell(row=ri, column=2, value=val).font = DATA_FONT

    ws2 = wb.create_sheet('Jahresübersicht')
    ws2['A1'] = 'Produkt'
    ws2['B1'] = 'Gesamtumsatz'
    ws2['A1'].font = HEADER_FONT
    ws2['B1'].font = HEADER_FONT
    ws2['A1'].fill = ws2['B1'].fill = HEADER_FILL
    ws2['A2'] = 'Konsolidieren Sie Q1+Q2+Q3 mit: Daten → Konsolidieren'

    # Update instructions on main sheet
    for r in range(3, 20):
        try:
            ws.unmerge_cells(f'A{r}:F{r}')
        except ValueError:
            pass
    ws['A1'] = 'Übung 5.3 — Konsolidierung'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79')
    write_instructions(ws, [
        'Wechseln Sie zum Blatt „Jahresübersicht".',
        'Konsolidieren Sie die drei Quartale Q1, Q2, Q3.',
        'Daten → Datentools → Konsolidieren → Funktion: Summe.',
        'Fügen Sie die Bereiche aus Q1, Q2 und Q3 hinzu.',
    ], 3)
    save(wb, 'M5_3_Konsolidierung')


def mod5_4():
    wb, ws = create_workbook('Übung 5.4 — Daten importieren')
    write_instructions(ws, [
        '1. Erstellen Sie eine CSV-Datei mit folgendem Inhalt:',
        '   Name,Alter,Stadt',
        '   Anna,28,Berlin',
        '   Tom,35,München',
        '   Lisa,22,Hamburg',
        '2. Importieren Sie diese Datei: Daten → Aus Text/CSV.',
        '3. Aktualisieren Sie die Verbindung nach einer CSV-Änderung.'
    ])
    save(wb, 'M5_4_Import')


# ============================================================
# MODUL 6
# ============================================================

def mod6_1_suchen():
    wb, ws = create_workbook('Übung 6.1 — Suchen und Ersetzen')
    headers = ['Name', 'Abteilung', 'Stadt', 'Eintrittsdatum']
    data = [
        ['Anna', 'Vertrieb', 'Berlin', '15.03.2020'],
        ['Tom', 'IT', 'München', '01.06.2019'],
        ['Lisa', 'Marketing', 'München', '10.01.2021'],
        ['Max', 'IT', 'Hamburg', '20.09.2018'],
        ['Sara', 'HR', 'München', '05.04.2022'],
        ['Paul', 'Vertrieb', 'Berlin', '12.11.2023'],
    ]
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        '1. Suchen Sie mit Strg+F alle Vorkommen von „München".',
        '2. Ersetzen Sie mit Strg+H alle „München" durch „München (Zentrale)".',
        '3. Suchen Sie mit der Option „Ganze Zellinhalte" nach „500" und',
        '   beobachten Sie den Unterschied zur normalen Suche.',
    ])
    save(wb, 'M6_1_Suchen_Ersetzen')


def mod6_2_fenster():
    wb, ws = create_workbook('Übung 6.2 — Fenster einfrieren')
    headers = ['Monat', 'Umsatz', 'Kosten', 'Gewinn', 'Marge %']
    data = [
        ['Januar', 5000, 3200, None, None],
        ['Februar', 4800, 3100, None, None],
        ['März', 5200, 3300, None, None],
        ['April', 6100, 3500, None, None],
        ['Mai', 5800, 3400, None, None],
        ['Juni', 6300, 3600, None, None],
        ['Juli', 5900, 3450, None, None],
        ['August', 6200, 3550, None, None],
    ]
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        '1. Fixieren Sie die oberste Zeile (Ansicht → Fenster einfrieren).',
        '2. Scrollen Sie nach unten — die Kopfzeile sollte sichtbar bleiben.',
        '3. Heben Sie die Fixierung auf und fixieren Sie dann Zeile 1 UND Spalte A.',
        '4. Scrollen Sie diagonal — Zeile 1 und Spalte A bleiben fixiert.',
    ])
    save(wb, 'M6_2_Fenster_fixieren')


def mod6_1():
    wb, ws = create_workbook('Übung 6.1 — Daten sortieren')
    headers = ['Name', 'Abteilung', 'Eintrittsdatum', 'Gehalt']
    data = [
        ['Anna', 'Vertrieb', '15.03.2020', 45000],
        ['Tom', 'IT', '01.06.2019', 52000],
        ['Lisa', 'Vertrieb', '10.01.2021', 42000],
        ['Max', 'IT', '20.09.2018', 55000],
        ['Sara', 'HR', '05.04.2022', 40000],
    ]
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        '1. Sortieren Sie alphabetisch nach Name.',
        '2. Sortieren Sie zuerst nach Abteilung, dann nach Gehalt (absteigend).',
        '3. Sortieren Sie nach Eintrittsdatum (älteste zuerst).'
    ])
    save(wb, 'M6_3_Sortieren')


def mod6_2():
    wb, ws = create_workbook('Übung 6.2 — Daten filtern')
    headers = ['Name', 'Abteilung', 'Eintrittsdatum', 'Gehalt']
    data = [
        ['Anna', 'Vertrieb', '15.03.2020', 45000],
        ['Tom', 'IT', '01.06.2019', 52000],
        ['Lisa', 'Vertrieb', '10.01.2021', 42000],
        ['Max', 'IT', '20.09.2018', 55000],
        ['Sara', 'HR', '05.04.2022', 40000],
    ]
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        'Aktivieren Sie den AutoFilter (Strg+Shift+L) und filtern Sie:',
        '1. Nur Mitarbeiter der IT-Abteilung.',
        '2. Alle Gehälter über 45.000 €.',
        '3. Mitarbeiter, die nach dem 01.01.2020 eingetreten sind.',
        '4. Kombination: IT-Abteilung UND Gehalt > 50.000 €.'
    ])
    save(wb, 'M6_4_Filtern')


def mod6_3():
    wb, ws = create_workbook('Übung 6.3 — Excel-Tabellen')
    headers = ['Name', 'Abteilung', 'Eintrittsdatum', 'Gehalt']
    data = [
        ['Anna', 'Vertrieb', '15.03.2020', 45000],
        ['Tom', 'IT', '01.06.2019', 52000],
        ['Lisa', 'Vertrieb', '10.01.2021', 42000],
        ['Max', 'IT', '20.09.2018', 55000],
        ['Sara', 'HR', '05.04.2022', 40000],
    ]
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        '1. Wandeln Sie den Bereich in eine Tabelle um: Bereich markieren → Strg+T.',
        '2. Benennen Sie die Tabelle: „Mitarbeiter" (Tabellenentwurf).',
        '3. Fügen Sie eine Ergebniszeile hinzu → Summe der Gehälter.',
        '4. Neue Zeile anhängen — Tabelle erweitert sich automatisch.'
    ])
    save(wb, 'M6_5_Tabellen')


def mod6_4():
    wb, ws = create_workbook('Übung 6.4 — Teilergebnisse')
    headers = ['Name', 'Abteilung', 'Gehalt']
    data = [
        ['Anna', 'Vertrieb', 45000],
        ['Lisa', 'Vertrieb', 42000],
        ['Paul', 'Vertrieb', 48000],
        ['Tom', 'IT', 52000],
        ['Max', 'IT', 55000],
        ['Uwe', 'IT', 51000],
        ['Sara', 'HR', 40000],
        ['Mia', 'HR', 43000],
        ['Kai', 'HR', 41000],
    ]
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        '1. Sortieren Sie zuerst nach Abteilung.',
        '2. Berechnen Sie Teilergebnisse: Summe Gehalt pro Abteilung.',
        '3. Nutzen Sie die Gliederungssymbole (1, 2, 3) zum Ein-/Ausblenden.'
    ])
    save(wb, 'M6_6_Teilergebnisse')


# ============================================================
# MODUL 7
# ============================================================

def mod7_1():
    wb, ws = create_workbook('Übung 7.1 — Bedingte Summen')
    headers = ['Verkäufer', 'Region', 'Produkt', 'Umsatz']
    data = [
        ['Anna', 'Nord', 'A', 500],
        ['Tom', 'Süd', 'B', 300],
        ['Anna', 'Nord', 'B', 700],
        ['Lisa', 'Ost', 'A', 400],
        ['Tom', 'Süd', 'A', 600],
        ['Anna', 'Nord', 'A', 200],
    ]
    row = write_table(ws, headers, data, 7) + 1
    add_result_row(ws, 'Umsatz Anna:', '=SUMMEWENN(A8:A13;"Anna";D8:D13)', row)
    add_result_row(ws, 'Umsatz Region Nord:', '=SUMMEWENN(B8:B13;"Nord";D8:D13)', row + 1)
    add_result_row(ws, 'Anzahl Produkt A:', '=ZÄHLENWENN(C8:C13;"A")', row + 2)
    add_result_row(ws, 'Umsatz Anna+Nord:', '=SUMMEWENNS(D8:D13;A8:A13;"Anna";B8:B13;"Nord")', row + 3)
    add_result_row(ws, '⌀ Umsatz Süd:', '=MITTELWERTWENN(B8:B13;"Süd";D8:D13)', row + 4)
    save(wb, 'M7_1_Bedingte_Summen')


def mod7_2():
    wb, ws = create_workbook('Übung 7.2 — SVERWEIS')

    # Preistabelle
    ws['A1'] = 'Preistabelle'
    ws['A1'].font = Font(bold=True, size=12)
    price_headers = ['Produktcode', 'Produktname', 'Preis']
    price_data = [
        ['P001', 'Tastatur', 29.99],
        ['P002', 'Maus', 19.99],
        ['P003', 'Monitor', 199.99],
        ['P004', 'Drucker', 89.99],
    ]
    write_table(ws, price_headers, price_data, 3)

    start = 9
    ws.cell(row=start, column=1, value='Bestellungen').font = Font(bold=True, size=12)
    order_headers = ['Bestell-Nr', 'Produktcode', 'Menge', 'Produktname', 'Einzelpreis', 'Gesamtpreis']
    order_data = [
        ['B001', 'P003', 2, None, None, None],
        ['B002', 'P001', 5, None, None, None],
    ]
    write_table(ws, order_headers, order_data, start + 1)

    write_instructions(ws, [
        '1. Verwenden Sie SVERWEIS für Produktname: =SVERWEIS(B11;A4:C7;2;FALSCH)',
        '2. Verwenden Sie SVERWEIS für Einzelpreis: =SVERWEIS(B11;A4:C7;3;FALSCH)',
        '3. Gesamtpreis = Menge × Einzelpreis',
        '4. Optional: WENNFEHLER für nicht gefundene Produktcodes'
    ], start + 5)
    save(wb, 'M7_2_SVERWEIS')


def mod7_3():
    wb, ws = create_workbook('Übung 7.3 — INDEX + VERGLEICH')
    headers = ['Produktcode', 'Produktname', 'Preis']
    data = [
        ['P001', 'Tastatur', 29.99],
        ['P002', 'Maus', 19.99],
        ['P003', 'Monitor', 199.99],
        ['P004', 'Drucker', 89.99],
    ]
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        '1. Finden Sie den Preis von „Monitor" mit INDEX+VERGLEICH:',
        '   =INDEX(C8:C11;VERGLEICH("Monitor";B8:B11;0))',
        '2. Erstellen Sie eine Formel, die vom Produktnamen zum Code sucht (nach links!):',
        '   =INDEX(A8:A11;VERGLEICH("Drucker";B8:B11;0))',
        '3. Warum ist INDEX+VERGLEICH flexibler als SVERWEIS?'
    ])
    save(wb, 'M7_3_INDEX_VERGLEICH')


def mod7_4():
    wb, ws = create_workbook('Übung 7.4 — Text- und Datumsfunktionen')
    headers = ['Vollständiger Name', 'Vorname', 'Nachname', 'E-Mail']
    data = [
        ['Max Mustermann', None, None, None],
        ['Anna Schmidt', None, None, None],
        ['Tom Meyer', None, None, None],
    ]
    row = write_table(ws, headers, data, 7)
    write_instructions(ws, [
        '1. Vorname: =LINKS(A8;FINDEN(" ";A8)-1)',
        '2. Nachname: =RECHTS(A8;LÄNGE(A8)-FINDEN(" ";A8))',
        '3. E-Mail: =KLEIN(B8)&"."&KLEIN(C8)&"@firma.de"',
        '4. Alter aus Geburtsdatum: =(HEUTE()-GebDat)/365 (als Zahl formatieren)'
    ])
    save(wb, 'M7_4_Text_Datum')


# ============================================================
# MODUL 8
# ============================================================

def mod8_1():
    wb, ws = create_workbook('Übung 8.1 — Erste Diagramme')
    headers = ['Monat', 'Umsatz']
    data = [
        ['Jan', 12000],
        ['Feb', 15000],
        ['Mär', 11000],
        ['Apr', 18000],
        ['Mai', 16000],
        ['Jun', 20000],
    ]
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        '1. Markieren Sie die Daten und erstellen Sie ein Säulendiagramm.',
        '2. Erstellen Sie ein Liniendiagramm für den Trend.',
        '3. Experimentieren Sie mit dem Diagrammtyp-Wechsel.'
    ])
    save(wb, 'M8_1_Erste_Diagramme')


def mod8_2():
    wb, ws = create_workbook('Übung 8.2 — Diagramme formatieren')
    headers = ['Monat', 'Umsatz']
    data = [
        ['Jan', 12000],
        ['Feb', 15000],
        ['Mär', 11000],
        ['Apr', 18000],
        ['Mai', 16000],
        ['Jun', 20000],
    ]
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        'Erstellen Sie ein Säulendiagramm und formatieren Sie:',
        '1. Diagrammtitel: „Monatsumsatz 2026"',
        '2. Achsen: X = „Monat", Y = „Umsatz (€)"',
        '3. Datenbeschriftungen zu den Säulen hinzufügen.',
        '4. Farbe der Säulen auf einheitliches Blau ändern.',
        '5. Legende entfernen (nur eine Datenreihe).'
    ])
    save(wb, 'M8_2_Diagrammformat')


def mod8_3():
    wb, ws = create_workbook('Übung 8.3 — Verbunddiagramm')
    headers = ['Monat', 'Umsatz', 'Gewinn']
    data = [
        ['Jan', 12000, 3000],
        ['Feb', 15000, 4500],
        ['Mär', 11000, 2500],
        ['Apr', 18000, 6000],
        ['Mai', 16000, 5000],
        ['Jun', 20000, 7000],
    ]
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        '1. Verbunddiagramm: Umsatz als Säulen, Gewinn als Linie (2. Achse).',
        '2. Sparklines in Spalte D für Umsatz-Trend (Einfügen → Sparklines → Linie).',
        '3. Sparklines in Spalte E für monatliche Veränderung (Typ: Gewinn/Verlust).'
    ])
    save(wb, 'M8_3_Verbunddiagramm')


def mod8_4():
    wb, ws = create_workbook('Übung 8.4 — Mini-Dashboard')
    headers = ['Produkt', 'Umsatz', 'Ziel', 'Zielerreichung %']
    data = [
        ['Produkt A', 15000, 18000, None],
        ['Produkt B', 22000, 20000, None],
        ['Produkt C', 8000, 12000, None],
        ['Produkt D', 19000, 18000, None],
    ]
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        '1. Zielerreichung: =B8/C8 (als % formatieren).',
        '2. Bedingte Formatierung: Farbskalen auf Umsatz-Spalte.',
        '3. Ampelsymbole auf Zielerreichung: >90%=grün, 70-90%=gelb, <70%=rot.',
        '4. Kreisdiagramm für Umsatzverteilung nach Produkt.',
        '5. Liniendiagramm für Monatstrend (Daten in weiterem Blatt).'
    ])
    save(wb, 'M8_4_Dashboard')


# ============================================================
# MODUL 9
# ============================================================

def mod9_1():
    wb, ws = create_workbook('Übung 9.1 — Erste Pivot-Tabelle')
    headers = ['Datum', 'Verkäufer', 'Region', 'Produkt', 'Kategorie', 'Umsatz', 'Menge']
    data = [
        ['01.01.2026', 'Anna', 'Nord', 'A', 'Hardware', 500, 2],
        ['15.01.2026', 'Tom', 'Süd', 'B', 'Software', 300, 1],
        ['03.02.2026', 'Anna', 'Nord', 'B', 'Software', 700, 3],
        ['18.02.2026', 'Lisa', 'Ost', 'A', 'Hardware', 400, 2],
        ['05.03.2026', 'Tom', 'Süd', 'A', 'Hardware', 600, 2],
        ['20.03.2026', 'Anna', 'Nord', 'A', 'Hardware', 200, 1],
        ['02.04.2026', 'Lisa', 'Ost', 'B', 'Software', 350, 2],
        ['15.04.2026', 'Tom', 'Süd', 'B', 'Software', 450, 2],
        ['01.05.2026', 'Anna', 'Nord', 'C', 'Zubehör', 800, 4],
        ['18.05.2026', 'Lisa', 'Ost', 'C', 'Zubehör', 250, 1],
        ['03.06.2026', 'Tom', 'Süd', 'A', 'Hardware', 550, 2],
        ['20.06.2026', 'Anna', 'Nord', 'B', 'Software', 900, 3],
        ['05.07.2026', 'Lisa', 'Ost', 'A', 'Hardware', 480, 2],
        ['18.07.2026', 'Tom', 'Süd', 'C', 'Zubehör', 320, 1],
        ['02.08.2026', 'Anna', 'Nord', 'A', 'Hardware', 670, 3],
        ['15.08.2026', 'Lisa', 'Ost', 'B', 'Software', 530, 2],
        ['01.09.2026', 'Tom', 'Süd', 'A', 'Hardware', 410, 2],
        ['18.09.2026', 'Anna', 'Nord', 'C', 'Zubehör', 290, 1],
        ['03.10.2026', 'Lisa', 'Ost', 'A', 'Hardware', 750, 3],
        ['20.10.2026', 'Tom', 'Süd', 'B', 'Software', 620, 2],
    ]
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        '1. Bereich markieren → Einfügen → PivotTable → Neues Blatt.',
        '2. Zeilen: Region, Spalten: Kategorie, Werte: Umsatz (Summe).',
        '3. Variieren: Verkäufer in Zeilen, Produkt in Spalten.',
        '4. Filter für Region hinzufügen. Daten aktualisieren testen.'
    ])
    save(wb, 'M9_1_Pivot')


def mod9_2():
    wb, ws = create_workbook('Übung 9.2 — Pivot anpassen')
    # Same data as 9.1
    headers = ['Datum', 'Verkäufer', 'Region', 'Produkt', 'Kategorie', 'Umsatz', 'Menge']
    data = [
        ['01.01.2026', 'Anna', 'Nord', 'A', 'Hardware', 500, 2],
        ['15.01.2026', 'Tom', 'Süd', 'B', 'Software', 300, 1],
        ['03.02.2026', 'Anna', 'Nord', 'B', 'Software', 700, 3],
        ['18.02.2026', 'Lisa', 'Ost', 'A', 'Hardware', 400, 2],
        ['05.03.2026', 'Tom', 'Süd', 'A', 'Hardware', 600, 2],
        ['20.03.2026', 'Anna', 'Nord', 'A', 'Hardware', 200, 1],
        ['02.04.2026', 'Lisa', 'Ost', 'B', 'Software', 350, 2],
        ['15.04.2026', 'Tom', 'Süd', 'B', 'Software', 450, 2],
        ['01.05.2026', 'Anna', 'Nord', 'C', 'Zubehör', 800, 4],
        ['18.05.2026', 'Lisa', 'Ost', 'C', 'Zubehör', 250, 1],
    ]
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        '1. Pivot-Tabelle erstellen.',
        '2. Umsatz als % des Gesamtergebnisses anzeigen.',
        '3. Daten nach Monaten gruppieren.',
        '4. Berechnetes Feld „Durchschnittspreis" = Umsatz / Menge.',
        '5. Zusammenfassung für Menge auf Mittelwert ändern.'
    ])
    save(wb, 'M9_2_Pivot_Anpassung')


def mod9_3():
    wb, ws = create_workbook('Übung 9.3 — Datenschnitte')
    headers = ['Datum', 'Verkäufer', 'Region', 'Produkt', 'Kategorie', 'Umsatz', 'Menge']
    data = [
        ['01.01.2026', 'Anna', 'Nord', 'A', 'Hardware', 500, 2],
        ['15.01.2026', 'Tom', 'Süd', 'B', 'Software', 300, 1],
        ['03.02.2026', 'Anna', 'Nord', 'B', 'Software', 700, 3],
        ['18.02.2026', 'Lisa', 'Ost', 'A', 'Hardware', 400, 2],
        ['05.03.2026', 'Tom', 'Süd', 'A', 'Hardware', 600, 2],
        ['20.03.2026', 'Anna', 'Nord', 'A', 'Hardware', 200, 1],
    ]
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        '1. Pivot-Tabelle erstellen.',
        '2. Datenschnitte für „Region" und „Kategorie" einfügen.',
        '3. Zeitachse für „Datum" hinzufügen.',
        '4. Zwei Pivot-Tabellen mit demselben Datenschnitt verbinden.'
    ])
    save(wb, 'M9_3_Slicer')


def mod9_4():
    wb, ws = create_workbook('Übung 9.4 — Interaktiver Bericht')
    headers = ['Datum', 'Verkäufer', 'Region', 'Produkt', 'Kategorie', 'Umsatz', 'Menge']
    data = [
        ['01.01.2026', 'Anna', 'Nord', 'A', 'Hardware', 500, 2],
        ['15.01.2026', 'Tom', 'Süd', 'B', 'Software', 300, 1],
        ['03.02.2026', 'Anna', 'Nord', 'B', 'Software', 700, 3],
        ['18.02.2026', 'Lisa', 'Ost', 'A', 'Hardware', 400, 2],
        ['05.03.2026', 'Tom', 'Süd', 'A', 'Hardware', 600, 2],
        ['20.03.2026', 'Anna', 'Nord', 'A', 'Hardware', 200, 1],
    ]
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        'Erstellen Sie ein interaktives Dashboard:',
        '1. Pivot-Tabelle: Umsatz nach Region und Kategorie.',
        '2. PivotChart (Säulendiagramm) daneben.',
        '3. Datenschnitte für Region und Jahr.',
        '4. Test: Wenn Sie einen Slicer-Wert ändern, aktualisiert sich alles.'
    ])
    save(wb, 'M9_4_PivotChart')


# ============================================================
# MODUL 10
# ============================================================

def mod10_1():
    wb, ws = create_workbook('Übung 10.1 — Zielwertsuche')
    ws.unmerge_cells('A1:F1')

    ws['A3'] = 'Kreditrechner'
    ws['A3'].font = Font(bold=True, size=14, color='1F4E79')

    labels = ['Kreditbetrag:', 'Zinssatz (p.a.):', 'Laufzeit (Jahre):', 'Monatliche Rate:']
    values = [200000, 0.04, 20, None]
    for i, (lbl, val) in enumerate(zip(labels, values), 5):
        ws.cell(row=i, column=1, value=lbl).font = Font(bold=True)
        cell = ws.cell(row=i, column=2, value=val)
        cell.font = DATA_FONT
        if i == 8:
            cell.value = '=RMZ(B6/12;B7*12;-B5)'
            cell.font = Font(name='Consolas', size=11, color='1F4E79', bold=True)

    write_instructions(ws, [
        '1. Die monatliche Rate wird mit =RMZ(B6/12;B7*12;-B5) berechnet.',
        '2. Zielwertsuche: Welcher Kreditbetrag ergibt max. 1.000 € Rate?',
        '   → Daten → Was-wäre-wenn → Zielwertsuche',
        '   → Zielzelle: B8, Zielwert: 1000, Veränderbare Zelle: B5',
        '3. Szenario-Manager: Vergleichen Sie 3%, 4% und 5% Zinsen.',
    ], 10)
    save(wb, 'M10_1_Zielwertsuche')


def mod10_2():
    wb, ws = create_workbook('Übung 10.2 — Finanzfunktionen')
    ws.unmerge_cells('A1:F1')
    ws['A3'] = 'Finanzberechnungen'
    ws['A3'].font = Font(bold=True, size=14, color='1F4E79')

    # Darlehen
    ws['A5'] = '1. Darlehen (15.000 €, 6%, 5 Jahre)'
    ws['A5'].font = Font(bold=True)
    ws['A6'] = 'Monatliche Rate:'
    ws['B6'] = '=RMZ(6%/12;5*12;-15000)'
    ws['B6'].font = Font(name='Consolas', color='1F4E79')

    # Sparen
    ws['A8'] = '2. Sparen (200 €/Monat, 5%, 30 Jahre)'
    ws['A8'].font = Font(bold=True)
    ws['A9'] = 'Endkapital:'
    ws['B9'] = '=ZW(5%/12;30*12;-200)'
    ws['B9'].font = Font(name='Consolas', color='1F4E79')

    # Investition
    ws['A11'] = '3. Investition (50.000 €, Rückflüsse)'
    ws['A11'].font = Font(bold=True)
    for ci, h in enumerate(['Jahr', 'Rückfluss'], 1):
        ws.cell(row=12, column=ci, value=h).font = HEADER_FONT
        ws.cell(row=12, column=ci).fill = HEADER_FILL
    flows = [0, 12000, 15000, 18000, 15000, 10000]
    for i, v in enumerate(flows):
        ws.cell(row=13 + i, column=1, value=i)
        ws.cell(row=13 + i, column=2, value=v)
    ws['A20'] = 'Kapitalwert (NBW):'
    ws['B20'] = '=NBW(8%;B13:B17)-50000'
    ws['B20'].font = Font(name='Consolas', color='1F4E79', bold=True)

    save(wb, 'M10_2_Finanzfunktionen')


# ============================================================
# MODUL 11
# ============================================================


def mod10_3():
    wb, ws = create_workbook('Übung 10.3 — Datentabelle')
    write_instructions(ws, [
        '1. Eindimensionale Datentabelle: Berechnen Sie die RMZ-Rate für',
        '   Zinssätze von 2% bis 8% (in 0,5%-Schritten).',
        '   Kredit: 250.000 €, Laufzeit: 30 Jahre.',
        '   Markieren Sie den Bereich und wählen Sie:',
        '   Daten → Was-wäre-wenn-Analyse → Datentabelle.',
        '2. Zweidimensionale Datentabelle: RMZ-Rate für Zinssätze (3%-7%)',
        '   als Zeilen-Eingabezelle und Laufzeiten (10-30 Jahre) als',
        '   Spalten-Eingabezelle.',
    ])
    save(wb, 'M10_3_Datentabelle')


def mod10_4():
    wb, ws = create_workbook('Übung 10.4 — Integrierte Finanzanalyse')
    write_instructions(ws, [
        'Investition: 500.000 €, jährliche Rückflüsse: 80.000 €, Dauer: 10 Jahre.',
        '1. Berechnen Sie NBW bei 6% Zinssatz.',
        '2. Zielwertsuche: Welcher Zinssatz ergibt NBW = 0? (=IKV)',
        '3. Datentabelle: NBW für Zinssätze 2%-12% (1%-Schritte).',
        '4. Ab welchem Zinssatz wird NBW negativ?',
    ])
    save(wb, 'M10_4_Finanzanalyse')


def mod11_1():
    wb, ws = create_workbook('Übung 11.1 — Drucklayout')
    headers = ['Jan', 'Feb', 'Mär', 'Apr', 'Mai', 'Jun', 'Jul', 'Aug']
    data = [[12000, 15000, 11000, 18000, 16000, 20000, 14000, 17000]]
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        '1. Ausrichtung auf Querformat ändern.',
        '2. Tabelle auf 1 Seite Breite skalieren.',
        '3. Daten horizontal und vertikal zentrieren.',
        '4. In der Seitenansicht (Strg+F2) überprüfen.'
    ])
    save(wb, 'M11_1_Drucklayout')


def mod11_2():
    wb, ws = create_workbook('Übung 11.2 — Druckbereich')
    headers = ['ID', 'Name', 'Wert']
    data = []
    for i in range(1, 51):
        data.append([i, f'Eintrag {i}', i * 100])
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        '1. Druckbereich festlegen: ersten 10 Zeilen markieren → Seite einrichten → Druckbereich.',
        '2. Manuellen Seitenumbruch nach Zeile 25 einfügen.',
        '3. Seitenumbruchvorschau prüfen.'
    ])
    save(wb, 'M11_2_Druckbereich')


def mod11_3():
    wb, ws = create_workbook('Übung 11.3 — Kopf-/Fußzeilen')
    headers = ['Monat', 'Umsatz', 'Kosten', 'Gewinn']
    data = [
        ['Januar', 50000, 32000, None],
        ['Februar', 48000, 31000, None],
        ['März', 52000, 33000, None],
    ]
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        '1. Kopfzeile: links = Firmenname, Mitte = „Quartalsbericht Q1 2026", rechts = Datum.',
        '2. Fußzeile: „Seite [Seite] von [Seiten]" + Dateipfad.',
        '3. Erste Zeile (Titel) auf jeder Seite wiederholen.',
        '   → Seite einrichten → Blatt → Wiederholungszeilen oben.'
    ])
    save(wb, 'M11_3_Kopfzeilen')


def mod11_4():
    wb, ws = create_workbook('Übung 11.4 — Zusammenarbeit')
    headers = ['Aufgabe', 'Verantwortlich', 'Status', 'Fällig']
    data = [
        ['Bericht erstellen', 'Anna', 'In Arbeit', '15.04.2026'],
        ['Daten prüfen', 'Tom', 'Offen', '20.04.2026'],
        ['Präsentation', 'Lisa', 'Erledigt', '10.04.2026'],
    ]
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        '1. Fügen Sie Kommentare zu Zellen hinzu (Überprüfen → Neuer Kommentar).',
        '2. Aktivieren Sie „Änderungen nachverfolgen".',
        '3. Exportieren Sie die Tabelle als PDF.'
    ])
    save(wb, 'M11_4_Zusammenarbeit')


# ============================================================
# MODUL 12
# ============================================================

def mod12_1():
    wb, ws = create_workbook('Übung 12.1 — Schutz')
    headers = ['Name', 'Gehalt', 'Bonus', 'Gesamt']
    data = [
        ['Anna', 45000, None, None],
        ['Tom', 52000, None, None],
        ['Lisa', 42000, None, None],
    ]
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        '1. Bonus (C8:C10) und Gesamt (D8:D10) sind Formelzellen zum Ausfüllen.',
        '2. Entsperren Sie die Eingabezellen (B8:B10): Zellen formatieren → Schutz → Gesperrt abwählen.',
        '3. Blattschutz aktivieren mit Kennwort „excel123".',
        '4. Testen: Kann man gesperrte Zellen bearbeiten?',
        '5. Schutz wieder aufheben.'
    ])
    save(wb, 'M12_1_Schutz')


def mod12_2():
    wb, ws = create_workbook('Übung 12.2 — Tastenkombinationen')
    headers = ['Aufgabe', 'Tastenkombination', 'Erledigt?']
    shortcuts = [
        ['Zum Rand springen', 'Strg+Pfeiltaste', '☐'],
        ['AutoFilter', 'Strg+Shift+L', '☐'],
        ['AutoSumme', 'Alt+=', '☐'],
        ['Wiederholen / Bezug', 'F4', '☐'],
        ['Zellen formatieren', 'Strg+1', '☐'],
        ['Spalte markieren', 'Strg+Leertaste', '☐'],
        ['Zeile markieren', 'Shift+Leertaste', '☐'],
        ['Zeilen einfügen', 'Strg+Shift++', '☐'],
    ]
    write_table(ws, headers, shortcuts, 7)
    write_instructions(ws, [
        'Führen Sie jede Tastenkombination mindestens 3× aus und haken Sie ab.'
    ])
    save(wb, 'M12_2_Tastenkombinationen')


def mod12_3():
    wb, ws = create_workbook('Übung 12.3 — Dokumentinspektion')
    write_instructions(ws, [
        '1. Datei → Informationen → Auf Probleme überprüfen → Dokument prüfen.',
        '2. Entfernen Sie alle gefundenen persönlichen Informationen.',
        '3. Speichern Sie eine Kopie im .xlsx-Format für Kompatibilität.',
        '✅ Verfahrensübung ohne vordefinierte Daten.'
    ])
    save(wb, 'M12_3_Inspektion')


# ============================================================
# MODUL 13
# ============================================================

def mod13_1():
    wb, ws = create_workbook('Übung 13.1 — Entwicklertools')
    write_instructions(ws, [
        '1. Aktivieren Sie die Registerkarte „Entwicklertools".',
        '   Datei → Optionen → Menüband anpassen → Entwicklertools ✅',
        '2. Überprüfen Sie die Makrosicherheitseinstellungen.',
        '3. Speichern Sie die Arbeitsmappe als MakroTest.xlsm.',
        '✅ Einrichtungsübung ohne vordefinierte Daten.'
    ])
    save(wb, 'M13_1_Entwicklertools')


def mod13_2():
    wb, ws = create_workbook('Übung 13.2 — Makro aufzeichnen')
    headers = ['Produkt', 'Preis', 'Menge', 'Gesamt']
    data = [
        ['A', 10, 5, None],
        ['B', 20, 3, None],
        ['C', 15, 8, None],
    ]
    write_table(ws, headers, data, 7)
    write_instructions(ws, [
        'Nehmen Sie ein Makro auf, das folgende Formatierung durchführt:',
        '1. Kopfzeile: Fett, zentriert, blaue Füllfarbe (#1F4E79), weiße Schrift.',
        '2. Außenrahmen um die gesamte Tabelle.',
        '3. Spaltenbreite automatisch anpassen.',
        'Speichern Sie als .xlsm. Führen Sie das Makro auf Blatt2 aus.'
    ])
    save(wb, 'M13_2_Makro_Aufzeichnen')


def mod13_3():
    wb, ws = create_workbook('Übung 13.3 — VBA-Editor')
    write_instructions(ws, [
        '1. Öffnen Sie Modul 13.2 (.xlsm) und drücken Sie Alt+F11.',
        '2. Suchen Sie das aufgezeichnete Makro im VBA-Editor.',
        '3. Ändern Sie die Schriftfarbe von Weiß auf Hellgelb (RGB: 255,255,200).',
        '4. Fügen Sie eine Meldung hinzu: MsgBox "Formatierung abgeschlossen!"',
        '5. Führen Sie das bearbeitete Makro aus.'
    ])
    save(wb, 'M13_3_VBA_Editor')


def mod13_4():
    wb, ws = create_workbook('Übung 13.4 — VBA programmieren')
    # Create multiple sheets for the VBA exercise
    for name in ['Januar', 'Februar', 'März']:
        if name != 'Übung':
            s = wb.create_sheet(name)
            s['A2'] = f'Daten für {name}'

    ws = wb['Übung']
    write_instructions(ws, [
        'Schreiben Sie ein Makro (Alt+F11), das:',
        '1. Alle Blätter der Arbeitsmappe durchläuft (For Each-Schleife).',
        '2. Auf jedem Blatt in Zelle A1 den Blattnamen einträgt.',
        '3. Eine Meldung ausgibt: „X Blätter verarbeitet!"',
        '',
        'Code-Vorlage:',
        'Sub BlattnamenEintragen()',
        '    Dim ws As Worksheet',
        '    Dim anzahl As Integer: anzahl = 0',
        '    For Each ws In ThisWorkbook.Worksheets',
        '        ws.Range("A1").Value = ws.Name',
        '        anzahl = anzahl + 1',
        '    Next ws',
        '    MsgBox anzahl & " Blätter verarbeitet!"',
        'End Sub'
    ])
    save(wb, 'M13_4_VBA_Programmieren')


def save(wb, name):
    path = os.path.join(OUT_DIR, f'{name}.xlsx')
    wb.save(path)
    print(f'  ✅ {name}.xlsx')


# ============================================================
# GENERATE ALL
# ============================================================
if __name__ == '__main__':
    print(f'Generiere Übungsdateien in: {OUT_DIR}\n')

    print('Modul 1: Einführung...')
    mod1_1(); mod1_2(); mod1_3(); mod1_4()

    print('Modul 2: Dateneingabe...')
    mod2_1(); mod2_2(); mod2_3(); mod2_4()

    print('Modul 3: Formatierung...')
    mod3_1(); mod3_2(); mod3_3(); mod3_4()

    print('Modul 4: Formeln...')
    mod4_1(); mod4_2(); mod4_3_namen(); mod4_3(); mod4_4()

    print('Modul 5: Validierung...')
    mod5_1(); mod5_2(); mod5_3(); mod5_4()

    print('Modul 6: Tabellen & Filter...')
    mod6_1_suchen(); mod6_2_fenster(); mod6_1(); mod6_2(); mod6_3(); mod6_4()

    print('Modul 7: Erweiterte Funktionen...')
    mod7_1(); mod7_2(); mod7_3(); mod7_4()

    print('Modul 8: Diagramme...')
    mod8_1(); mod8_2(); mod8_3(); mod8_4()

    print('Modul 9: Pivot-Tabellen...')
    mod9_1(); mod9_2(); mod9_3(); mod9_4()

    print('Modul 10: Finanzfunktionen...')
    mod10_1(); mod10_2()

    print('Modul 10: Analyse...')
    mod10_3(); mod10_4()

    print('Modul 11: Druck...')
    mod11_1(); mod11_2(); mod11_3(); mod11_4()

    print('Modul 12: Schutz...')
    mod12_1(); mod12_2(); mod12_3()

    print('Modul 13: Makros...')
    mod13_1(); mod13_2(); mod13_3(); mod13_4()

    print(f'\n✅ Fertig! {sum(1 for _ in os.listdir(OUT_DIR) if _.endswith(".xlsx"))} Dateien in {OUT_DIR}')
